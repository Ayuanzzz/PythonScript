import openpyxl

# 指定Excel文件路径
excel_file = r"D:\Python\test\xls\贡则寺.xlsx"
sheet_name = '宗教活动场所建（构）筑物统计表'  # 选择第二个工作表
skip_rows = 3  # 跳过前两行

excel_file2 = r"D:\Python\test\xls\建（构）筑物.xlsx"
sheet_name2 = '建（构）筑物模板'

# 打开Excel文件
workbook = openpyxl.load_workbook(excel_file, data_only=True)
worksheet = workbook[sheet_name]

workbook2 = openpyxl.load_workbook(excel_file2)
worksheet2 = workbook2[sheet_name2]

# 初始化一个空的字典
a = {}

# 遍历工作表中的每一行，从第三行开始
for row_index in range(skip_rows, worksheet.max_row + 1):
    row = worksheet[row_index]
    # 使用A和B列作为元组的键，将C列数据和行号添加到值列表中
    key = (float(row[5].value), float(row[6].value))
    value = row[12].value  # 行号从1开始

    if key in a:
        print("重复")
        a[key].append(value)
    else:
        a[key] = value

print(f'a---------------{a}')

# 获取要修改的工作表
# output_sheet = workbook2[sheet_name2]

for row_index in range(skip_rows, worksheet2.max_row + 1):
    row = worksheet2[row_index]

    key = (float(row[7].value), float(row[8].value))
    value = row_index

    if key in a:
        # 修改单元格的值
        col_index = 2  # 修改第二列（列号从1开始）
        new_value = a[key]
        worksheet2.cell(row=row_index, column=col_index, value=new_value)

# 保存修改后的 Excel 文件
# output_file = r"D:\Python\test\xls\建（构）筑物nnnn.xlsx"
workbook2.save(excel_file2 )

print("\nDone")
