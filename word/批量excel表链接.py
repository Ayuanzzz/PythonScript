import openpyxl
import os

# 指定要搜索的文件夹路径
folder_path = r'D:\Python\test\袁伟'

def joinExcel(excel_file, excel_file2):
    print('------------------------------------------------')
    print(excel_file)
    print(excel_file2)
    print('------------------------------------------------')
    # 指定Excel文件路径
    # excel_file = r"D:\Python\test\xls\贡则寺.xlsx"
    sheet_name = '宗教活动场所建（构）筑物统计表'  # 选择第二个工作表
    skip_rows = 3  # 跳过前两行
    # excel_file2 = r"D:\Python\test\xls\建（构）筑物.xlsx"
    sheet_name2 = '建（构）筑物模板'
    # 打开Excel文件
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    worksheet = workbook[sheet_name]
    workbook2 = openpyxl.load_workbook(excel_file2)
    worksheet2 = workbook2[sheet_name2]
    # 初始化一个空的字典
    a = {}
    # 遍历工作表中的每一行，从第三行开始
    for row_index in range(skip_rows, worksheet.max_row + 1):
        row = worksheet[row_index]
        # 使用A和B列作为元组的键，将C列数据和行号添加到值列表中
        if not row[5].value or not row[6].value:
            continue
        key = (float(row[5].value), float(row[6].value))
        value = row[12].value  # 行号从1开始
        if key not in a:
            a[key] = value

    for row_index in range(skip_rows, worksheet2.max_row + 1):
        row = worksheet2[row_index]
        if not row[7].value or not row[8].value:
            continue
        key = (float(row[7].value), float(row[8].value))
        if key in a:
            # 修改单元格的值
            col_index = 2  # 修改第二列（列号从1开始）
            new_value = a[key]
            worksheet2.cell(row=row_index, column=col_index, value=new_value)
    workbook2.save(excel_file2)



xlsx_list = []
# 使用os.listdir()列出文件夹中的所有文件
for root, dirs, files in os.walk(folder_path):
    for file in files:
        if file == '建（构）筑物.xlsx':
            file_path = os.path.join(root, file)
            xlsx_list.append(file_path)


def findExcel(file_path):
    # 使用os.path.split()函数分割文件路径
    directories = file_path.split(os.path.sep)
    # 获取倒数第二层目录
    if len(directories) >= 2:
        second_last_directory = directories[-2]
        temple_list = second_last_directory.split("基础")
        temple_name = temple_list[0]

    table_path = r'D:\Python\test\2121'
    # 使用os.listdir()列出文件夹中的所有文件和子文件夹
    for root, dirs, files in os.walk(table_path):
        for file in files:
            file_name, file_extension = os.path.splitext(file)
            if file_name == temple_name:
                print(file_name)
                excel_path = os.path.join(root, file)
                try:
                    # print(excel_path)
                    # print(file_path)
                    joinExcel(excel_path, file_path)
                except Exception as e:
                    # 处理其他异常的代码
                    print(f'err------{file_path}')
                    print(f"发生了异常：{e}")


for file_path in xlsx_list:
    print(file_path)
    findExcel(file_path)

print("\nDone")
