import openpyxl
import os

def toTxt(excel_path):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(excel_path)  # 替换为你的 Excel 文件路径
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]  # 替换为你的工作表名称
    print(sheet)
    # 打开文本文件以写入模式
    txt_path = excel_path.replace('.xlsx', '.txt')
    with open(txt_path, 'w') as txt_file:  # 输出的文本文件名
        # 遍历工作表的每一行，从第三行开始
        for i, row in enumerate(sheet.iter_rows(), start=1):
            if i > 2:  # 忽略前两行
                # 遍历行中的每个单元格，并将其值写入文本文件，用制表符分隔
                row_data = '\t'.join([str(cell.value) for cell in row]) + '\n'
                txt_file.write(row_data)

# 指定文件夹路径
folder_path = input("folder: ")  # 替换为你的文件夹路径

# 获取文件夹中所有文件的文件名
file_names = os.listdir(folder_path)

# 筛选出扩展名为 .txt 的文件路径
txt_file_paths = [os.path.join(folder_path, file) for file in file_names if file.endswith('.xlsx')]

# 打印所有 .txt 文件的路径
for txt_path in txt_file_paths:
    toTxt(txt_path)

print(f'Excel 内容已写入文本文件')

input("exit")
